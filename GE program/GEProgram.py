import openpyxl
import os
import sys
import time
import pandas as pd

def progress(count, total, status=''):
    bar_len = 60
    filled_len = int(round(bar_len * count / float(total)))

    percents = round(100.0 * count / float(total), 2)
    bar = '=' * filled_len + '-' * (bar_len - filled_len)
    if(percents.is_integer()):
        sys.stdout.write('[%s] %s%s ...%s\r' % (bar, percents, '%', status))
        sys.stdout.flush()
        print()
        
def progress2(count, total, status=''):
    bar_len = 60
    filled_len = int(round(bar_len * count / float(total)))

    percents = round(100.0 * count / float(total), 3)
    bar = '=' * filled_len + '-' * (bar_len - filled_len)

    sys.stdout.write('[%s] %s%s ...%s\r' % (bar, percents, '%', status))
    sys.stdout.flush()

wb = openpyxl.load_workbook('AH+SS - Copy.xlsx')
sheet = wb.active
i = 4
while i <= 9:
    sheet.delete_cols(4)
    i = i + 1
sheet.delete_cols(12)
sheet.delete_cols(10)
sheet.delete_cols(9)
sheet.delete_cols(5)
sheet.insert_cols(4)

total = sheet.max_row
del_row = []
for row in sheet.iter_rows(min_row=1, min_col = 1, max_col = 11):
    progress(row[1].row, total)
    for cell in row:
        if(cell.value == "Prereq:"):
            sheet.cell(row=cell.row - 1,column= 4).value = sheet.cell(row=cell.row, column=cell.column+1).value
            del_row.append(cell.row)
        elif(cell.value == "AH" or cell.value == "SS" or cell.value == "ACGH" or cell.value == "DD" or cell.value == "SL" or cell.value == "WC" or cell.value == "WE"):
            cell.value = "x"

         

print("Deleting Duplicates....\n")
time.sleep(1)
total2 = sheet.max_row
values = []
for row1 in sheet.iter_rows(min_row=1, max_col = 1):
    for cell in row1:
        progress(cell.row, total2)
        if cell.value in values:
            del_row.append(cell.row)
        elif cell.value == None :
            del_row.append(cell.row)
        else:
            values.append(cell.value)

print("Deleting Rows....\n")
del_row = list(set(del_row))
del_row.sort(reverse=True)

for value in del_row:
    progress2(del_row.index(value), len(del_row))
    sheet.delete_rows(value)

    
sheet.insert_rows(1)
sheet.cell(row=1, column = 1).value = "Course"
sheet.cell(row=1, column = 2).value = "Title"
sheet.cell(row=1, column = 3).value = "Units"
sheet.cell(row=1, column = 5).value = "AH"
sheet.cell(row=1, column = 6).value = "SS"
sheet.cell(row=1, column = 7).value = "ACGH"
sheet.cell(row=1, column = 8).value = "DD"
sheet.cell(row=1, column = 9).value = "SL"
sheet.cell(row=1, column = 10).value = "WC"
sheet.cell(row=1, column = 11).value = "WE"

wb.save('AH+SS - Copy.xlsx')
print('Sorting completed!')

df = pd.read_excel('AH+SS - Copy.xlsx')
df = df.sort_values(by='Course')
df.to_excel('Master GE List.xlsx', index = False)

wb1 = openpyxl.load_workbook('Master GE List.xlsx')
sheet1 = wb1.active

sheet1.insert_rows(1)
sheet1.insert_rows(1)

sheet1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
sheet1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
sheet1.merge_cells(start_row= 2, start_column=4, end_row=3, end_column=4)
sheet1.merge_cells(start_row=2, start_column=5, end_row=2, end_column=6)
sheet1.merge_cells(start_row=2, start_column=7, end_row=2, end_column=11)

sheet1.cell(row=1, column = 1).value = "GE Master List (geared towards needed ECE GE credits)"
sheet1.cell(row=2, column = 1).value = "Course Information"
sheet1.cell(row=2, column = 4).value = "Required prerequisites & Course Limitations"
sheet1.cell(row=2, column = 5).value = "Breadth"
sheet1.cell(row=2, column = 7).value = "Core Literacies"

wb1.save('Master GE List.xlsx')
print('Formatting completed!')



